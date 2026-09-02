"""DOL OFLC LCA disclosure ingestion (§6d): stream-parse the quarterly XLSX into lca_wages.

Kept: certified, full-time, computer-occupation SOC codes (15-12xx), annualized offered wage,
prevailing wage level. ~250 MB per quarterly file; openpyxl read-only streaming, one pass.
Caveats encoded in the prior: Level I/II only for new-grad lookups; the filed wage is a floor.
"""

from __future__ import annotations

import logging
import re
import sqlite3
from pathlib import Path
from typing import Any

import httpx

from radar import db
from radar.config import Config
from radar.fetch.registry import normalize_company_name
from radar.parse.locations import parse_location

log = logging.getLogger("radar.lca")

PERF_PAGE = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
SOC_PREFIXES = (
    "15-12",
    "15-11",
    "15-13",
    "15-20",
)  # software devs, computer/info research, analysts, math/ops research
UNIT_MULT = {"Year": 1.0, "Hour": 2080.0, "Week": 52.0, "Bi-Weekly": 26.0, "Month": 12.0}


def discover_latest_url(client: httpx.Client) -> tuple[str, str]:
    """Return (fiscal_tag, absolute_url) of the newest LCA disclosure file on the performance page."""
    r = client.get(PERF_PAGE)
    r.raise_for_status()
    found = re.findall(r'href="([^"]*LCA_Disclosure_Data_(FY\d{4}_Q\d)\.xlsx)"', r.text)
    if not found:
        raise RuntimeError("no LCA disclosure links found on the DOL performance page")
    found.sort(key=lambda t: t[1], reverse=True)
    href, tag = found[0]
    url = href if href.startswith("http") else "https://www.dol.gov" + href
    return tag, url.replace("//media", "/media")


def download(url: str, dest: Path, client: httpx.Client) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 1_000_000:
        return dest
    tmp = dest.with_suffix(".part")
    with client.stream("GET", url) as r, tmp.open("wb") as f:
        r.raise_for_status()
        for chunk in r.iter_bytes(1 << 20):
            f.write(chunk)
    tmp.rename(dest)
    return dest


def _annual(value: Any, unit: Any) -> float | None:
    try:
        v = float(str(value).replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return None
    mult = UNIT_MULT.get(str(unit or "Year").strip().title())
    if mult is None:
        return None
    return v * mult


def ingest_file(
    conn: sqlite3.Connection, path: Path, fiscal_tag: str, *, progress: Any = None
) -> dict[str, int]:
    from openpyxl import load_workbook

    if db.scalar(conn, "SELECT COUNT(*) FROM lca_wages WHERE fiscal_file = ?", (fiscal_tag,)):
        return {"skipped": 1, "reason": "already ingested"}  # type: ignore[dict-item]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h or "").strip().upper() for h in next(rows)]
    col = {name: i for i, name in enumerate(header)}

    def g(row: tuple, name: str) -> Any:
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    kept = seen = 0
    batch: list[tuple] = []
    metro_cache: dict[str, str | None] = {}
    insert_sql = (
        "INSERT INTO lca_wages (fiscal_file, employer_name, employer_norm, job_title, soc_code, worksite_city, "
        "worksite_state, metro, wage_level, wage_annual, prevailing_annual, decision_date) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)"
    )

    def flush() -> None:
        # short transactions: other writers (the fetch pipeline) must never be starved
        if batch:
            with db.transaction(conn):
                conn.executemany(insert_sql, batch)
            batch.clear()

    for row in rows:
        seen += 1
        if progress and seen % 50000 == 0:
            progress(seen, kept)
        if str(g(row, "CASE_STATUS") or "").strip().lower() != "certified":
            continue
        soc = str(g(row, "SOC_CODE") or "").strip()
        if not soc.startswith(SOC_PREFIXES):
            continue
        if str(g(row, "FULL_TIME_POSITION") or "Y").strip().upper() != "Y":
            continue
        wage = _annual(g(row, "WAGE_RATE_OF_PAY_FROM"), g(row, "WAGE_UNIT_OF_PAY"))
        if wage is None or wage < 20_000 or wage > 1_500_000:
            continue
        pw = _annual(g(row, "PREVAILING_WAGE"), g(row, "PW_UNIT_OF_PAY"))
        emp = str(g(row, "EMPLOYER_NAME") or "").strip()
        city = str(g(row, "WORKSITE_CITY") or "").strip()
        state = str(g(row, "WORKSITE_STATE") or "").strip()
        key = f"{city}, {state}"
        if key not in metro_cache:
            li = parse_location(key)
            metro_cache[key] = li.metro or (li.kind if li.kind in ("remote",) else None)
        lvl = str(g(row, "PW_WAGE_LEVEL") or "").strip().upper() or None
        dd = g(row, "DECISION_DATE")
        batch.append(
            (
                fiscal_tag,
                emp,
                normalize_company_name(emp),
                str(g(row, "JOB_TITLE") or "")[:120],
                soc,
                city[:60],
                state[:2],
                metro_cache[key],
                lvl,
                wage,
                pw,
                str(dd)[:10] if dd else None,
            )
        )
        kept += 1
        if len(batch) >= 5000:
            flush()
    flush()
    wb.close()
    return {"rows_seen": seen, "rows_kept": kept}


def refresh(
    conn: sqlite3.Connection, cfg: Config, *, fiscal: str | None = None, progress: Any = None
) -> dict[str, Any]:
    with httpx.Client(
        headers={"User-Agent": cfg.fetch.user_agent}, timeout=120, follow_redirects=True
    ) as client:
        tag, url = discover_latest_url(client)
        if fiscal:
            tag = fiscal
            url = f"https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_{fiscal}.xlsx"
        dest = cfg.data_dir / "lca" / f"LCA_Disclosure_Data_{tag}.xlsx"
        download(url, dest, client)
    stats = ingest_file(conn, dest, tag, progress=progress)
    db.kv_set(conn, "lca_last_refresh", {"fiscal": tag, "url": url, **stats})
    return {"fiscal": tag, "url": url, **stats}


class LcaPrior:
    """employer × SOC × metro lookup at Level I/II, held in memory (one query at construction).
    Returns {low, high, point, n} or None. Lookups are cached per (company, metro)."""

    def __init__(self, conn: sqlite3.Connection) -> None:
        import bisect

        self.conn = conn
        self.by_emp: dict[str, list[tuple[str | None, float]]] = {}
        self.loaded = False
        self._cache: dict[tuple, dict[str, Any] | None] = {}
        try:
            rows = db.all_rows(
                conn,
                "SELECT employer_norm, metro, wage_annual FROM lca_wages WHERE soc_code LIKE '15-12%' AND wage_level IN ('I','II')",
            )
        except sqlite3.OperationalError:  # table not migrated yet
            rows = []
        for r in rows:
            self.by_emp.setdefault(r["employer_norm"], []).append(
                (r["metro"], float(r["wage_annual"]))
            )
        self.keys = sorted(self.by_emp)
        self._bisect = bisect
        self.loaded = bool(self.by_emp)

    def _employers(self, norms: set[str]) -> list[str]:
        out: list[str] = []
        for n in norms:
            if n in self.by_emp:
                out.append(n)
            prefix = n + " "
            i = self._bisect.bisect_left(self.keys, prefix)
            while i < len(self.keys) and self.keys[i].startswith(prefix):
                out.append(self.keys[i])
                i += 1
        return out

    def lookup(self, company: dict[str, Any] | None, p: dict[str, Any]) -> dict[str, Any] | None:
        if not self.loaded or not company:
            return None
        metro = p.get("primary_metro")
        key = (company.get("id") or company.get("name"), metro)
        if key in self._cache:
            return self._cache[key]
        names = [company.get("name"), *(company.get("aliases") or [])]
        norms = {normalize_company_name(n) for n in names if n}
        norms = {n for n in norms if len(n) >= 4}
        res = None
        if norms:
            emps = self._employers(norms)
            wages = [w for e in emps for (m, w) in self.by_emp[e] if m == metro]
            if len(wages) < 3:
                wages = [w for e in emps for (_, w) in self.by_emp[e]]
            if len(wages) >= 3:
                wages.sort()
                n = len(wages)
                res = {
                    "low": wages[n // 4],
                    "high": wages[(3 * n) // 4],
                    "point": wages[n // 2],
                    "n": n,
                }
        self._cache[key] = res
        return res
